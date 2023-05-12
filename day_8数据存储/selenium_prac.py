from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os

browser = webdriver.Chrome()
browser.get('https://www.bilibili.com/')

wait = WebDriverWait(browser, 10)


def parse_html():
    result = []
    html = browser.page_source
    print('开始解析:' + browser.current_url)
    soup = BeautifulSoup(html, 'lxml')
    video_items = soup.find_all(attrs={'class': 'bili-video-card'})
    for video_item in video_items:
        tittle = video_item.find('h3', attrs={'class': 'bili-video-card__info--tit'}).text
        link = video_item.a.attrs['href']
        view = video_item.find('span', attrs={'class': 'bili-video-card__stats--item'}).span.text
        duration = video_item.find('span', attrs={'class': 'bili-video-card__stats__duration'}).text
        date = video_item.find('span', attrs={'class': 'bili-video-card__info--date'}).text
        result.append({'tittle': tittle,
                       'link': link,
                       'view': view,
                       'duration': duration,
                       'date': date})

    if os.path.exists('result.xlsx'):
        wb = load_workbook('result.xlsx')
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['tittle', 'link', 'view', 'duration', 'date'])  # 添加表头

    # 将数据写入工作表
    for row in result:
        ws.append([row['tittle'], row['link'], row['view'], row['duration'], row['date']])

    # 保存工作簿
    wb.save('result.xlsx')


input = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'nav-search-input')))
submit = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'nav-search-btn')))

input.send_keys('蔡徐坤 篮球')
submit.click()

print('跳转到新窗口')
all_h = browser.window_handles
browser.switch_to.window(all_h[1])
parse_html()

total = wait.until(
    EC.presence_of_element_located((By.XPATH, '//button[contains(@class,"vui_pagenation--btn-num")][last()]')))
print(total.text)

for i in range(2, int(total.text) + 1):
    print('获取下一页数据')
    next_bt = wait.until(
        EC.presence_of_element_located((By.XPATH, '//button[contains(@class,"vui_pagenation--btn-side")][last()]')))
    next_bt.click()
    wait.until(EC.text_to_be_present_in_element((By.CLASS_NAME, 'vui_button--active-blue'), str(i)))
    parse_html()
